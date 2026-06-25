# -*- coding: utf-8 -*-
"""
h5_to_radial_v2.py

代码首次修改于2026年5月14日
功能
----
1. 支持输入单个 h5 文件、日期文件夹，或日期文件夹的父文件夹（如年份文件夹）。
2. 保持原 h5_to_radial.py 的基本输出格式：
   - *_radial_wind_260514.npz
   - radial_wind_data_*_260514.xlsx
3. 从原始功率谱中计算 P/S 通道径向风速与 SNR：
   - ROI 仍采用频率 bin 82:164；
   - 有效距离门仍为第 4~60 门，对应输出 Col_1~Col_57；
   - 扣除首门噪声基底后，负值置 0；
   - 峰区范围采用“趋势反转峰区”；
   - 峰值位置采用趋势反转峰区内的谱质心；
   - SNR 分母采用“本径向、本通道首门 ROI 每-bin 噪声均值 × 当前门峰区 bin 数”；
   - 若峰区 bin 数相对于临近门中位数突变超过阈值，则该门判为无效。
4. 额外输出峰区积分值 Peak Sum 与归一化值 Peak Norm。
5. npz 中不保存 n_peak_bins / gate_valid 等 QC 中间字段；这些中间量仅额外写入 Excel，便于调试。

注意
----
- PEAK_METHOD 采用 'centroid'，这里的 centroid 指“趋势反转峰区内谱质心”。
- 不做固定 SNR 阈值判定；仅对无法识别峰区、峰区积分为 0、噪声分母为 0、峰区 bin 数突变等情况置 NaN。
"""

from __future__ import annotations

import os
from pathlib import Path
from datetime import datetime
import multiprocessing as mp
from typing import Iterable

from pip._internal.resolution import legacy
from torch._C import dtype
from tqdm import tqdm

import h5py
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.interpolate import interp1d

from peak_estimator import PeakFinder


# ============================= 常改动参数区 =============================
# 绝对路径；可以是：
# 1. 单个 .h5 文件；
# 2. 直接包含 .h5 文件的日期文件夹；
# 3. 包含多个日期文件夹的父文件夹（如年份文件夹）。
INPUT_PATH = r"E:\测风组实验数据\RawData\2026"

# 输出目录。
OUTPUT_DIR = r"F:\3220240787\Lidar_Simulation\wind_inversion\los_velocity_and_snr\year_2026"

# 输出文件名前缀。设为 None 时自动使用：
# - 单个 h5 文件：h5 文件名；
# - 日期文件夹：文件夹名；
# - 年份父文件夹：每个日期子文件夹名。
OUTPUT_STEM = None

# 若 INPUT_PATH 为年份父文件夹，是否递归查找日期子文件夹内的 h5。
# 通常设 False 即可：会查找每个直接子文件夹内的 *.h5。
RECURSIVE_SEARCH_IN_DATE_FOLDER = False

# 多进程数量。为 1 时不启用 multiprocessing；多个 h5 文件时可适当增大。
N_PROCESSES = 30

# 峰值检索方法：当前实现中 centroid 表示“趋势反转峰区内谱质心”。
PEAK_METHOD = "centroid"

# 是否额外保存 Excel，便于人工检查。
WRITE_EXCEL = True

# 是否使用centroid_method()的异常结果作为额外SNR掩膜
USE_LEGACY_CENTROID_VALID_MASK = True

# ROI 与距离门范围：保持原 h5_to_radial.py 的 82:164 与第4~60门。
ROI_SLICE = slice(82, 164)
KEEP_GATE_START = 3   # Python index；3 表示原始第4门
KEEP_GATE_END = 59    # Python index；59 表示原始第60门，包含

# 扣除首门噪声后，负值是否置为 0。
CLIP_NEGATIVE_AFTER_NOISE_SUBTRACTION = True

# 峰区 bin 数突变判据：与前后各 PEAK_BIN_JUMP_WINDOW 门的局部中位数比较。
PEAK_BIN_JUMP_WINDOW = 2
PEAK_BIN_JUMP_THRESHOLD = 2.0

# SNR 计算中的极小正值只用于判断“是否为 0 或非法”，不用于强行除以 eps 得到 Inf/极大值。
SNR_DENOM_EPS = 1e-12

# 归一化参数。
P_LO_P_W = 1.85e-3
P_LO_S_W = 1.64e-3
E_PULSE_J = 40e-6
N_ACC = 10000
# =======================================================================


def read_h5(filename: str | Path):
    """
    读取单个 h5 文件。

    返回
    ----
    P_corr, S_corr : ndarray
        形状为 [径向, 60门, ROI频率bin] 的扣噪后功率谱；第4门及以后已扣除首门噪声，
        且可选负值置 0。第1~3门保留原始 ROI 值，主要用于保持维度一致。
    P_raw_roi, S_raw_roi : ndarray
        未扣噪的 ROI 功率谱，用于计算首门噪声均值。
    azimuthAngle, timeStamp : ndarray
        原始方位角和时间。
    """
    filename = Path(filename)
    with h5py.File(filename, "r") as file_T:
        spec_raw = file_T["specData"][:file_T["specData"].shape[0] - 1, :]
        specData = spec_raw.reshape(spec_raw.shape[0], 120, 512).astype(np.float64)

        azimuthAngle = file_T["azimuthAngle"][:file_T["specData"].shape[0] - 1].flatten()
        azimuthAngle[np.argwhere(azimuthAngle >= 360)[:, 0]] = \
            azimuthAngle[np.argwhere(azimuthAngle >= 360)[:, 0]] - 360

        timeStamp_raw = file_T["timeStamp"][:file_T["specData"].shape[0] - 1]
        timeStamp = np.array([
            datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
            for ts in timeStamp_raw.flatten()
        ])

    front_60 = specData[:, 0:60, :]
    back_60 = specData[:, 60:, :]

    P_raw_roi = front_60[:, :, ROI_SLICE].copy()
    S_raw_roi = back_60[:, :, ROI_SLICE].copy()

    P_corr = P_raw_roi.copy()
    S_corr = S_raw_roi.copy()

    P_corr[:, KEEP_GATE_START:, :] -= P_raw_roi[:, 0:1, :]
    S_corr[:, KEEP_GATE_START:, :] -= S_raw_roi[:, 0:1, :]

    if CLIP_NEGATIVE_AFTER_NOISE_SUBTRACTION:
        P_corr[:, KEEP_GATE_START:, :] = np.maximum(P_corr[:, KEEP_GATE_START:, :], 0.0)
        S_corr[:, KEEP_GATE_START:, :] = np.maximum(S_corr[:, KEEP_GATE_START:, :], 0.0)

    return P_corr, S_corr, P_raw_roi, S_raw_roi, azimuthAngle, timeStamp


def calculate_radial_velocity(index, aom_shift=120e6, wavelength=1550e-9):
    """
    根据 ROI 内峰值索引换算径向风速。
    """
    freqs = np.fft.fftfreq(1024, 1 / 1e9)[ROI_SLICE]
    interp_function = interp1d(
        np.arange(len(freqs)),
        freqs,
        kind="linear",
        fill_value="extrapolate",
    )
    interpolated_freqs = interp_function(index)
    delta_freqs = interpolated_freqs - aom_shift
    radial_velocity = -wavelength * delta_freqs / 2
    return radial_velocity


def find_peak_bounds_by_trend_change(spectrum_1d: np.ndarray):
    """
    趋势反转法识别峰区边界。

    返回
    ----
    peak_idx, left_idx, right_idx
        若无法识别有效峰区，返回 (-1, -1, -1)。
    """
    y = np.asarray(spectrum_1d, dtype=np.float64)
    if y.ndim != 1 or y.size == 0:
        return -1, -1, -1
    if np.all(~np.isfinite(y)):
        return -1, -1, -1

    y_max = np.nanmax(y)
    if (not np.isfinite(y_max)) or y_max <= 0:
        return -1, -1, -1

    peak_idx = int(np.nanargmax(y))
    n = y.size

    left_idx = 0
    for k in range(peak_idx - 1, 0, -1):
        if y[k - 1] > y[k]:
            left_idx = k
            break

    right_idx = n - 1
    for k in range(peak_idx + 1, n - 1):
        if y[k + 1] > y[k]:
            right_idx = k
            break

    if left_idx > peak_idx:
        left_idx = peak_idx
    if right_idx < peak_idx:
        right_idx = peak_idx

    return peak_idx, left_idx, right_idx


def compute_centroid_index_and_sum(spectrum_1d: np.ndarray, left_idx: int, right_idx: int):
    """
    在指定峰区内计算谱质心索引与峰区积分。
    """
    if left_idx < 0 or right_idx < 0 or right_idx < left_idx:
        return np.nan, np.nan, 0

    y = np.asarray(spectrum_1d, dtype=np.float64)
    segment = y[left_idx:right_idx + 1]
    peak_sum = np.sum(segment, dtype=np.float64)
    n_peak_bins = right_idx - left_idx + 1

    if (not np.isfinite(peak_sum)) or peak_sum <= 0:
        return np.nan, np.nan, n_peak_bins

    idx = np.arange(left_idx, right_idx + 1, dtype=np.float64)
    centroid_idx = np.sum(idx * segment, dtype=np.float64) / peak_sum
    return centroid_idx, peak_sum, n_peak_bins


def normalize_peak_sum(peak_sum: np.ndarray, lo_power_w: float):
    """
    对峰区积分值进行归一化。
    """
    return peak_sum / (lo_power_w * E_PULSE_J * float(N_ACC))


def detect_peak_bin_jump_invalid(n_peak_bins: np.ndarray):
    """
    检查峰区 bin 数是否相对临近距离门突变。

    判据：同一径向内，某门 n_peak_bins 与前后各 PEAK_BIN_JUMP_WINDOW 门的局部中位数
    差值 > PEAK_BIN_JUMP_THRESHOLD，则判为无效。

    不设置峰区 bin 数的绝对上下限。
    """
    n_peak_bins = np.asarray(n_peak_bins, dtype=float)
    n_radial, n_gate = n_peak_bins.shape
    invalid = np.zeros((n_radial, n_gate), dtype=bool)

    for i in range(n_radial):
        for j in range(n_gate):
            current = n_peak_bins[i, j]
            if (not np.isfinite(current)) or current <= 0:
                invalid[i, j] = True
                continue

            left = max(0, j - PEAK_BIN_JUMP_WINDOW)
            right = min(n_gate, j + PEAK_BIN_JUMP_WINDOW + 1)
            neighbor_idx = [k for k in range(left, right) if k != j]
            if not neighbor_idx:
                continue

            neighbor_values = n_peak_bins[i, neighbor_idx]
            neighbor_values = neighbor_values[np.isfinite(neighbor_values) & (neighbor_values > 0)]
            if neighbor_values.size == 0:
                continue

            local_median = np.nanmedian(neighbor_values)
            if np.abs(current - local_median) > PEAK_BIN_JUMP_THRESHOLD:
                invalid[i, j] = True

    return invalid


def retrieve_channel_products(spec_corr: np.ndarray, spec_raw_roi: np.ndarray, lo_power_w: float):
    """
    对单通道功率谱计算：峰区质心索引、径向风速、SNR、峰区 Sum、峰区归一化值、QC 信息。

    参数
    ----
    spec_corr : ndarray
        扣除首门噪声且负值置 0 后的 ROI 谱，形状 [径向, 60, ROI_bin]。
    spec_raw_roi : ndarray
        未扣噪 ROI 谱，形状 [径向, 60, ROI_bin]。
    lo_power_w : float
        对应通道本振功率。
    """
    n_radial = spec_corr.shape[0]
    n_output_gates = KEEP_GATE_END - KEEP_GATE_START + 1

    peak_index = np.full((n_radial, n_output_gates), np.nan, dtype=np.float64)
    peak_sum = np.full((n_radial, n_output_gates), np.nan, dtype=np.float64)
    n_peak_bins = np.zeros((n_radial, n_output_gates), dtype=np.float64)
    legacy_centroid_valid = np.ones((n_radial, n_output_gates), dtype=bool)

    for i in range(n_radial):
        for gate in range(KEEP_GATE_START, KEEP_GATE_END + 1):
            out_gate = gate - KEEP_GATE_START
            spectrum = spec_corr[i, gate, :]

            if USE_LEGACY_CENTROID_VALID_MASK:
                legacy_spectrum = spec_raw_roi[i, gate, :] - spec_raw_roi[i, 0, :]
                legacy_noise = spec_raw_roi[i, 0, :]

                try:
                    legacy_peak, _, _, legacy_snr = PeakFinder.centroid_method(
                    legacy_spectrum,
                    legacy_noise
                    )

                    if (not np.isfinite(legacy_peak)) or (not np.isfinite(legacy_snr)):
                        legacy_centroid_valid[i, out_gate] = False

                except Exception:
                    legacy_centroid_valid[i, out_gate] = False

            _, left_idx, right_idx = find_peak_bounds_by_trend_change(spectrum)
            centroid_idx, current_sum, current_bins = compute_centroid_index_and_sum(
                spectrum, left_idx, right_idx
            )
            peak_index[i, out_gate] = centroid_idx
            peak_sum[i, out_gate] = current_sum
            n_peak_bins[i, out_gate] = current_bins

    # 峰区 bin 数突变判定。
    # jump_invalid = detect_peak_bin_jump_invalid(n_peak_bins)

    # 每条径向使用自己的首门 ROI 每-bin 噪声均值。
    noise_mean_per_bin = np.nanmean(spec_raw_roi[:, 0, :], axis=1).astype(np.float64)
    noise_power_equiv = noise_mean_per_bin[:, None] * n_peak_bins

    valid = (
        np.isfinite(peak_index)
        & np.isfinite(peak_sum)
        & (peak_sum > 0)
        & np.isfinite(noise_power_equiv)
        & (noise_power_equiv > SNR_DENOM_EPS)
        # & (~jump_invalid)
        & legacy_centroid_valid
    )

    snr = np.full_like(peak_sum, np.nan, dtype=np.float64)
    snr[valid] = 10.0 * np.log10(peak_sum[valid] / noise_power_equiv[valid])

    peak_sum_out = peak_sum.copy()
    peak_sum_out[~valid] = np.nan

    peak_norm = normalize_peak_sum(peak_sum_out, lo_power_w)
    peak_norm[~valid] = np.nan

    peak_index_out = peak_index.copy()
    peak_index_out[~valid] = np.nan

    radial_velocity = calculate_radial_velocity(peak_index_out)
    radial_velocity[~valid] = np.nan

    return {
        "radial_velocity": radial_velocity.astype(np.float64),
        "snr": snr.astype(np.float64),
        "peak_sum": peak_sum_out.astype(np.float64),
        "peak_norm": peak_norm.astype(np.float64),
        "n_peak_bins": n_peak_bins.astype(np.float64),
        "gate_valid": valid,
    }


def process_file(filename: str | Path):
    """
    处理单个 h5 文件：趋势反转峰区 -> 谱质心径向风速 -> SNR -> Peak Sum/Norm。
    若文件损坏无法读取，返回 None 并打印警告。
    """
    try:
        P_corr, S_corr, P_raw_roi, S_raw_roi, azi_data, time = read_h5(filename)
    except OSError as e:
        print(f"[跳过] 文件损坏，无法打开：{filename}")
        print(f"       错误信息：{e}")
        return None

    if PEAK_METHOD.lower() != "centroid":
        raise ValueError("当前版本仅实现 PEAK_METHOD='centroid'，即趋势反转峰区内谱质心。")

    p_products = retrieve_channel_products(P_corr, P_raw_roi, P_LO_P_W)
    s_products = retrieve_channel_products(S_corr, S_raw_roi, P_LO_S_W)

    return (
        p_products["radial_velocity"],
        s_products["radial_velocity"],
        azi_data,
        time,
        p_products["snr"],
        s_products["snr"],
        p_products["peak_sum"],
        s_products["peak_sum"],
        p_products["peak_norm"],
        s_products["peak_norm"],
        p_products["n_peak_bins"],
        s_products["n_peak_bins"],
        p_products["gate_valid"],
        s_products["gate_valid"],
    )


def collect_h5_jobs(input_path: str | Path):
    """
    根据输入路径生成处理任务。

    返回
    ----
    jobs : list[tuple[str, list[Path]]]
        每个元素为 (output_stem, h5_files)。

    规则
    ----
    1. 输入为单个 h5：输出一个任务；
    2. 输入为直接包含 h5 的日期文件夹：输出一个任务；
    3. 输入为年份父文件夹：对其直接子文件夹逐个查找 h5，每个日期文件夹输出一个任务。
    """
    path = Path(input_path)
    if not path.is_absolute():
        raise ValueError(f"INPUT_PATH 必须是绝对路径：{input_path}")
    if not path.exists():
        raise FileNotFoundError(f"路径不存在：{input_path}")

    if path.is_file():
        if path.suffix.lower() != ".h5":
            raise ValueError(f"输入文件必须是 .h5 文件：{path}")
        stem = OUTPUT_STEM if OUTPUT_STEM else path.stem
        return [(stem, [path])]

    jobs: list[tuple[str, list[Path]]] = []

    # 情况1：该文件夹自身就是日期文件夹，直接包含 h5。
    direct_h5 = sorted(path.glob("*.h5"))
    if direct_h5:
        stem = OUTPUT_STEM if OUTPUT_STEM else path.name
        jobs.append((stem, direct_h5))

    # 情况2：该文件夹是父文件夹，直接子文件夹中包含 h5。
    for subdir in sorted([p for p in path.iterdir() if p.is_dir()]):
        pattern = "**/*.h5" if RECURSIVE_SEARCH_IN_DATE_FOLDER else "*.h5"
        h5_files = sorted(subdir.glob(pattern))
        if not h5_files:
            continue
        if OUTPUT_STEM and len(jobs) == 0:
            stem = f"{OUTPUT_STEM}_{subdir.name}"
        elif OUTPUT_STEM:
            stem = f"{OUTPUT_STEM}_{subdir.name}"
        else:
            stem = subdir.name
        jobs.append((stem, h5_files))

    if not jobs:
        raise FileNotFoundError(f"未在路径或其直接子文件夹中找到 h5 文件：{path}")

    return jobs


def make_output_dataframe(matrix: np.ndarray, time_all: np.ndarray, azi_data_all: np.ndarray):
    """
    按原 h5_to_radial.py 的 Excel 格式组织输出。
    """
    df = pd.DataFrame(matrix, columns=[f"Col_{i + 1}" for i in range(matrix.shape[1])])
    df.insert(0, "Azi_Data", azi_data_all.ravel())
    df.insert(0, "Time", time_all.ravel())
    return df


def save_outputs(results: list[tuple], output_dir: str | Path, output_stem: str, write_excel: bool = True):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    radial_v_P_all = np.concatenate([result[0] for result in results], axis=0)
    radial_v_S_all = np.concatenate([result[1] for result in results], axis=0)
    azi_data_all = np.concatenate([result[2] for result in results], axis=0)
    time_all = np.concatenate([result[3] for result in results], axis=0)
    SNR_P_all = np.concatenate([result[4] for result in results], axis=0)
    SNR_S_all = np.concatenate([result[5] for result in results], axis=0)
    peak_sum_P_all = np.concatenate([result[6] for result in results], axis=0)
    peak_sum_S_all = np.concatenate([result[7] for result in results], axis=0)
    peak_norm_P_all = np.concatenate([result[8] for result in results], axis=0)
    peak_norm_S_all = np.concatenate([result[9] for result in results], axis=0)

    # 仅 Excel 输出的 QC 中间量。
    n_peak_bins_P_all = np.concatenate([result[10] for result in results], axis=0)
    n_peak_bins_S_all = np.concatenate([result[11] for result in results], axis=0)
    gate_valid_P_all = np.concatenate([result[12] for result in results], axis=0)
    gate_valid_S_all = np.concatenate([result[13] for result in results], axis=0)

    npz_path = output_dir / f"{output_stem}_radial_wind_260514.npz"
    np.savez(
        npz_path,
        radial_v_P=radial_v_P_all,
        radial_v_S=radial_v_S_all,
        azi_data=azi_data_all,
        time=time_all,
        SNR_P=SNR_P_all,
        SNR_S=SNR_S_all,
        peak_sum_P=peak_sum_P_all,
        peak_sum_S=peak_sum_S_all,
        peak_norm_P=peak_norm_P_all,
        peak_norm_S=peak_norm_S_all,
    )
    print(f"已保存 NPZ：{npz_path}")

    if write_excel:
        excel_path = output_dir / f"radial_wind_data_{output_stem}_260514.xlsx"
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            make_output_dataframe(radial_v_P_all, time_all, azi_data_all).to_excel(
                writer, sheet_name="Radial_V_P", index=False
            )
            make_output_dataframe(radial_v_S_all, time_all, azi_data_all).to_excel(
                writer, sheet_name="Radial_V_S", index=False
            )
            make_output_dataframe(SNR_P_all, time_all, azi_data_all).to_excel(
                writer, sheet_name="SNR_P", index=False
            )
            make_output_dataframe(SNR_S_all, time_all, azi_data_all).to_excel(
                writer, sheet_name="SNR_S", index=False
            )
            make_output_dataframe(peak_sum_P_all, time_all, azi_data_all).to_excel(
                writer, sheet_name="Peak_Sum_P", index=False
            )
            make_output_dataframe(peak_sum_S_all, time_all, azi_data_all).to_excel(
                writer, sheet_name="Peak_Sum_S", index=False
            )
            make_output_dataframe(peak_norm_P_all, time_all, azi_data_all).to_excel(
                writer, sheet_name="Peak_Norm_P", index=False
            )
            make_output_dataframe(peak_norm_S_all, time_all, azi_data_all).to_excel(
                writer, sheet_name="Peak_Norm_S", index=False
            )

            # 额外 QC 中间量，后续若不需要可删除这部分输出。
            make_output_dataframe(n_peak_bins_P_all, time_all, azi_data_all).to_excel(
                writer, sheet_name="QC_N_Peak_Bins_P", index=False
            )
            make_output_dataframe(n_peak_bins_S_all, time_all, azi_data_all).to_excel(
                writer, sheet_name="QC_N_Peak_Bins_S", index=False
            )
            make_output_dataframe(gate_valid_P_all.astype(int), time_all, azi_data_all).to_excel(
                writer, sheet_name="QC_Gate_Valid_P", index=False
            )
            make_output_dataframe(gate_valid_S_all.astype(int), time_all, azi_data_all).to_excel(
                writer, sheet_name="QC_Gate_Valid_S", index=False
            )

        wb = load_workbook(excel_path)
        for sheet in wb.sheetnames:
            wb[sheet].sheet_state = "visible"
        wb.save(excel_path)
        print(f"已保存 Excel：{excel_path}")

    return npz_path


def run_one_job(output_stem: str, h5_files: list[Path]):
    print("=" * 60)
    print(f"输出标识：{output_stem}")
    print(f"h5 文件数：{len(h5_files)}")
    print("=" * 60)

    # --- 增量处理：输出已存在且比所有输入 h5 新则跳过 ---
    npz_path = Path(OUTPUT_DIR) / f"{output_stem}_radial_wind_260514.npz"
    if npz_path.exists():
        npz_mtime = npz_path.stat().st_mtime
        h5_mtimes = [h5.stat().st_mtime for h5 in h5_files]
        if h5_mtimes and all(npz_mtime >= m for m in h5_mtimes):
            print(f"[跳过] 输出已存在且为最新，无需重复处理：{npz_path}")
            return npz_path
        else:
            print(f"[增量] 检测到新增或更新的 h5 文件，将重新处理。")

    h5_file_strs = [str(h5_file) for h5_file in h5_files]

    if N_PROCESSES == 1 or len(h5_files) == 1:
        results = [
            process_file(h5_file)
            for h5_file in tqdm(
                h5_file_strs,
                desc=f"Processing {output_stem}",
                unit="file"
            )
        ]
    else:
        with mp.Pool(processes=N_PROCESSES) as pool:
            results = list(
                tqdm(
                    pool.imap(process_file, h5_file_strs),
                    total=len(h5_file_strs),
                    desc=f"Processing {output_stem}",
                    unit="file"
                )
            )

    # --- 过滤掉损坏文件返回的 None ---
    valid_results = [r for r in results if r is not None]
    skipped = len(results) - len(valid_results)
    if skipped > 0:
        print(f"[警告] {output_stem}：共跳过 {skipped} 个损坏文件。")

    if not valid_results:
        print(f"[错误] {output_stem} 下所有文件均无法处理，跳过输出。")
        return None

    return save_outputs(valid_results, OUTPUT_DIR, output_stem, write_excel=WRITE_EXCEL)


def main():
    jobs = collect_h5_jobs(INPUT_PATH)
    print(f"共发现 {len(jobs)} 个处理任务。")

    for output_stem, h5_files in jobs:
        run_one_job(output_stem, h5_files)


if __name__ == "__main__":
    main()
