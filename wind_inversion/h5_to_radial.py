import os
from pathlib import Path
import multiprocessing as mp

import h5py
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.interpolate import interp1d
from datetime import datetime

from peak_estimator import PeakFinder

# ============================= 常改动参数区 =============================
# 绝对路径；既可以是单个 .h5 文件，也可以是包含 .h5 文件的文件夹。
INPUT_PATH = r"E:\测风组实验数据\RawData\2026_05_20(Azimuth_test)\2026_05_20\22deg5"

# 输出目录。会保存 *_radial_wind_0325.npz 和可选 Excel。
OUTPUT_DIR = r"F:\3220240787\Lidar_Simulation\wind_inversion\inversion_results"

# 输出文件名前缀。设为 None 时：
# - 输入为单个 h5 文件：使用该 h5 文件名；
# - 输入为文件夹：使用文件夹名。
OUTPUT_STEM = None

# 文件夹输入时是否递归搜索子文件夹中的 h5 文件。
RECURSIVE_SEARCH = False

# 多进程数量。为 1 时不启用 multiprocessing；多个 h5 文件时可适当增大。
N_PROCESSES = 30

# 谱峰检索方法：
# 'gaussian'：沿用原代码，高斯拟合中心作为峰值索引；
# 'centroid'：谱质心法，centroid_method 返回的质心作为峰值索引。
# 注意：SNR 仍由 centroid_method 计算；若 PEAK_METHOD='centroid'，峰值索引和 SNR 来自同一次质心计算。
PEAK_METHOD = 'centroid'

# 是否额外保存 Excel，便于人工检查径向风速和 SNR。
WRITE_EXCEL = True
# =======================================================================


def read_h5(filename):
    """
    读取单个 h5 文件中的双通道功率谱数据，并沿用原代码的数据处理方式。

    注意：这里仍然保持原始算法思路：
    1. specData reshape 为 [径向数, 120个距离门, 512个频率bin]；
    2. 前 60 门作为 P 通道，后 60 门作为 S 通道；
    3. 截取频率 bin 82:164；
    4. 第 4 门及以后扣除第 1 门噪声基底。
    """
    file_T = h5py.File(filename, 'r')
    specData = file_T['specData'][:file_T['specData'].shape[0] - 1, :]\
        .reshape(file_T['specData'].shape[0] - 1, 120, 512)\
        .astype(np.float64)

    azimuthAngle = file_T['azimuthAngle'][:file_T['specData'].shape[0] - 1].flatten()
    azimuthAngle[np.argwhere(azimuthAngle >= 360)[:, 0]] = \
        azimuthAngle[np.argwhere(azimuthAngle >= 360)[:, 0]] - 360

    timeStamp = file_T['timeStamp'][:file_T['specData'].shape[0] - 1]
    timeStamp = np.array([datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')
                          for ts in timeStamp.flatten()])
    file_T.close()

    front_60 = specData[:, 0:60, :]
    back_60 = specData[:, 60:, :]
    front_60_select = front_60[:, :, 82:164].copy()
    back_60_select = back_60[:, :, 82:164].copy()

    front_60_select[:, 3:, :] -= front_60_select[:, 0:1, :]
    back_60_select[:, 3:, :] -= back_60_select[:, 0:1, :]

    noise_front = front_60_select[:, 0, :]
    noise_back = back_60_select[:, 0, :]

    return front_60_select, back_60_select, azimuthAngle, timeStamp, noise_front, noise_back


def calculate_radial_velocity(index, aom_shift=120e6, wavelength=1550e-9):
    """
    根据功率谱峰值索引换算径向风速。保持原代码公式不变。
    """
    freqs = np.fft.fftfreq(1024, 1 / 1e9)[82:164]
    interp_function = interp1d(np.arange(len(freqs)), freqs, kind='linear', fill_value='extrapolate')
    interpolated_freqs = interp_function(index)
    delta_freqs = interpolated_freqs - aom_shift
    radial_velocity = -wavelength * delta_freqs / 2
    return radial_velocity


def process_file(filename):
    """
    处理单个 h5 文件：功率谱寻峰 -> 径向风速 -> SNR。
    """
    print(filename)
    P_data, S_data, azi_data, time, noise_P, noise_S = read_h5(filename)

    index_P = np.zeros((P_data.shape[0], P_data.shape[1] - 3))
    index_S = np.zeros((S_data.shape[0], S_data.shape[1] - 3))
    SNR_P = np.zeros((P_data.shape[0], P_data.shape[1] - 3))
    SNR_S = np.zeros((S_data.shape[0], S_data.shape[1] - 3))

    for j in range(P_data.shape[0]):
        for i in range(3, P_data.shape[1]):
            if PEAK_METHOD.lower() == 'gaussian':
                peak_P, _, _ = PeakFinder.gaussian_fit(P_data[j, i, :], 1, noise_P[j, :])
                peak_S, _, _ = PeakFinder.gaussian_fit(S_data[j, i, :], 1, noise_S[j, :])
                _, _, _, snr_P = PeakFinder.centroid_method(P_data[j, i, :], noise_P[j, :])
                _, _, _, snr_S = PeakFinder.centroid_method(S_data[j, i, :], noise_S[j, :])
            elif PEAK_METHOD.lower() == 'centroid':
                peak_P, _, _, snr_P = PeakFinder.centroid_method(P_data[j, i, :], noise_P[j, :])
                peak_S, _, _, snr_S = PeakFinder.centroid_method(S_data[j, i, :], noise_S[j, :])
            else:
                raise ValueError(f"未知 PEAK_METHOD={PEAK_METHOD!r}，可选：'gaussian'、'centroid'。")

            index_P[j, i - 3] = peak_P
            index_S[j, i - 3] = peak_S
            SNR_P[j, i - 3] = snr_P
            SNR_S[j, i - 3] = snr_S

    radial_v_P = calculate_radial_velocity(index_P)
    radial_v_S = calculate_radial_velocity(index_S)

    return radial_v_P, radial_v_S, azi_data, time, SNR_P, SNR_S


def collect_h5_files(input_path, recursive=False):
    """
    从绝对路径中收集 h5 文件。input_path 可以是单个 h5 文件或文件夹。
    """
    path = Path(input_path)
    if not path.is_absolute():
        raise ValueError(f"INPUT_PATH 必须是绝对路径：{input_path}")
    if not path.exists():
        raise FileNotFoundError(f"路径不存在：{input_path}")

    if path.is_file():
        if path.suffix.lower() != '.h5':
            raise ValueError(f"输入文件必须是 .h5 文件：{path}")
        return [path]

    pattern = '**/*.h5' if recursive else '*.h5'
    h5_files = sorted(path.glob(pattern))
    if not h5_files:
        raise FileNotFoundError(f"未在路径中找到 h5 文件：{path}")
    return h5_files


def infer_output_stem(input_path, output_stem=None):
    if output_stem:
        return output_stem
    path = Path(input_path)
    return path.stem if path.is_file() else path.name


def save_outputs(results, output_dir, output_stem, write_excel=True):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    radial_v_P_all = np.concatenate([result[0] for result in results], axis=0)
    radial_v_S_all = np.concatenate([result[1] for result in results], axis=0)
    azi_data_all = np.concatenate([result[2] for result in results], axis=0)
    time_all = np.concatenate([result[3] for result in results], axis=0)
    SNR_P_all = np.concatenate([result[4] for result in results], axis=0)
    SNR_S_all = np.concatenate([result[5] for result in results], axis=0)

    npz_path = output_dir / f'{output_stem}_radial_wind_0506.npz'
    np.savez(npz_path,
             radial_v_P=radial_v_P_all,
             radial_v_S=radial_v_S_all,
             azi_data=azi_data_all,
             time=time_all,
             SNR_P=SNR_P_all,
             SNR_S=SNR_S_all)
    print(f"已保存 NPZ：{npz_path}")

    if write_excel:
        excel_path = output_dir / f"radial_wind_data_{output_stem}_0325.xlsx"
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df_radial_v_P = pd.DataFrame(radial_v_P_all,
                                         columns=[f"Col_{i + 1}" for i in range(radial_v_P_all.shape[1])])
            df_radial_v_P.insert(0, "Azi_Data", azi_data_all.ravel())
            df_radial_v_P.insert(0, "Time", time_all.ravel())
            df_radial_v_P.to_excel(writer, sheet_name="Radial_V_P", index=False)

            df_radial_v_S = pd.DataFrame(radial_v_S_all,
                                         columns=[f"Col_{i + 1}" for i in range(radial_v_S_all.shape[1])])
            df_radial_v_S.insert(0, "Azi_Data", azi_data_all.ravel())
            df_radial_v_S.insert(0, "Time", time_all.ravel())
            df_radial_v_S.to_excel(writer, sheet_name="Radial_V_S", index=False)

            df_SNR_P = pd.DataFrame(SNR_P_all, columns=[f"Col_{i + 1}" for i in range(SNR_P_all.shape[1])])
            df_SNR_P.insert(0, "Azi_Data", azi_data_all.ravel())
            df_SNR_P.insert(0, "Time", time_all.ravel())
            df_SNR_P.to_excel(writer, sheet_name="SNR_P", index=False)

            df_SNR_S = pd.DataFrame(SNR_S_all, columns=[f"Col_{i + 1}" for i in range(SNR_S_all.shape[1])])
            df_SNR_S.insert(0, "Azi_Data", azi_data_all.ravel())
            df_SNR_S.insert(0, "Time", time_all.ravel())
            df_SNR_S.to_excel(writer, sheet_name="SNR_S", index=False)

        wb = load_workbook(excel_path)
        for sheet in wb.sheetnames:
            wb[sheet].sheet_state = "visible"
        wb.save(excel_path)
        print(f"已保存 Excel：{excel_path}")

    return npz_path


def main():
    h5_files = collect_h5_files(INPUT_PATH, recursive=RECURSIVE_SEARCH)
    output_stem = infer_output_stem(INPUT_PATH, OUTPUT_STEM)

    if N_PROCESSES == 1 or len(h5_files) == 1:
        results = [process_file(str(h5_file)) for h5_file in h5_files]
    else:
        with mp.Pool(processes=N_PROCESSES) as pool:
            results = pool.map(process_file, [str(h5_file) for h5_file in h5_files])

    save_outputs(results, OUTPUT_DIR, output_stem, write_excel=WRITE_EXCEL)


if __name__ == '__main__':
    main()
